#!/usr/bin/env python3
"""Build the PMP-aligned Excel data template from the generated sample dataset."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

D = json.load(open("data/portfolio_data.json", encoding="utf-8"))

ARIAL = "Arial"
HDR_IN = PatternFill("solid", fgColor="1C5CAB")     # input columns
HDR_CALC = PatternFill("solid", fgColor="52514E")   # computed columns
HDR_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BODY = Font(name=ARIAL, size=10)
CALC = Font(name=ARIAL, size=10, color="000000")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN)

MONEY = '$#,##0;($#,##0);-'
PCT = '0.0%'
IDX = '0.00'
DATE = 'yyyy-mm-dd'

wb = Workbook()


def sheet(name, headers, rows, calc_from=None, widths=None, formats=None, freeze="A2"):
    ws = wb.create_sheet(name)
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HDR_FONT
        c.fill = HDR_CALC if (calc_from and j >= calc_from) else HDR_IN
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, r in enumerate(rows, 2):
        for j, h in enumerate(headers, 1):
            v = r.get(h)
            if v is None:
                continue
            c = ws.cell(row=i, column=j, value=v)
            c.font = BODY
            c.border = BORDER
            if formats and h in formats:
                c.number_format = formats[h]
    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 30
    for j, h in enumerate(headers, 1):
        w = (widths or {}).get(h, max(11, min(len(h) + 4, 30)))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    return ws


# ------------------------------------------------------------------ Priorities
sheet("Priorities",
      ["PriorityID", "PriorityName", "Description", "ExecutiveSponsor", "WeightPct", "FY"],
      [dict(r, WeightPct=r["WeightPct"]) for r in D["priorities"]],
      widths={"PriorityName": 34, "Description": 60, "ExecutiveSponsor": 18},
      formats={"WeightPct": PCT})

# ------------------------------------------------------------------ Portfolios
sheet("Portfolios",
      ["PortfolioID", "PortfolioName", "PriorityID", "PortfolioOwner", "Department"],
      D["portfolios"], widths={"PortfolioName": 30, "PortfolioOwner": 18})

# ------------------------------------------------------------------ Programs
sheet("Programs",
      ["ProgramID", "ProgramName", "PortfolioID", "ProgramManager"],
      D["programs"], widths={"ProgramName": 34, "ProgramManager": 20})

# ------------------------------------------------------------------ Projects
PJ_IN = ["ProjectID", "ProjectName", "ProgramID", "ProjectManager", "Department", "Phase",
         "Status", "StartDate", "BaselineFinish", "ForecastFinish", "PercentComplete",
         "BudgetAtCompletion", "PlannedValue", "EarnedValue", "ActualCost", "Sponsor", "BusinessCase"]
PJ_CALC = ["ScheduleVariance", "CostVariance", "SPI", "CPI", "EAC", "VAC", "SlipDays"]
ws = sheet("Projects", PJ_IN + PJ_CALC, D["projects"], calc_from=len(PJ_IN) + 1,
           widths={"ProjectName": 34, "ProjectManager": 18, "Phase": 24, "BusinessCase": 46,
                   "BudgetAtCompletion": 16, "PlannedValue": 14, "EarnedValue": 14,
                   "ActualCost": 14, "Sponsor": 16},
           formats={"PercentComplete": PCT, "BudgetAtCompletion": MONEY, "PlannedValue": MONEY,
                    "EarnedValue": MONEY, "ActualCost": MONEY,
                    "StartDate": DATE, "BaselineFinish": DATE, "ForecastFinish": DATE})

n = len(D["projects"]) + 1
for i in range(2, n + 1):
    ws[f"R{i}"] = f"=N{i}-M{i}"                              # SV = EV - PV
    ws[f"S{i}"] = f"=N{i}-O{i}"                              # CV = EV - AC
    ws[f"T{i}"] = f"=IFERROR(N{i}/M{i},\"\")"                # SPI = EV / PV
    ws[f"U{i}"] = f"=IFERROR(N{i}/O{i},\"\")"                # CPI = EV / AC
    ws[f"V{i}"] = f"=IFERROR(L{i}/(N{i}/O{i}),\"\")"         # EAC = BAC / CPI
    ws[f"W{i}"] = f"=IFERROR(L{i}-(L{i}/(N{i}/O{i})),\"\")"  # VAC = BAC - EAC
    ws[f"X{i}"] = f'=IFERROR(DATEDIF(I{i},J{i},"d"),0)'      # forecast slip vs baseline
    for col, fmt in (("R", MONEY), ("S", MONEY), ("T", IDX), ("U", IDX),
                     ("V", MONEY), ("W", MONEY), ("X", "0")):
        ws[f"{col}{i}"].font = CALC
        ws[f"{col}{i}"].number_format = fmt
        ws[f"{col}{i}"].border = BORDER

dv_phase = DataValidation(type="list", formula1='"0 - Requested,1 - Initiating,2 - Planning,3 - Executing,4 - Monitoring & Controlling,5 - Closing"', allow_blank=True)
dv_status = DataValidation(type="list", formula1='"Active,On Hold,Closed,Cancelled"', allow_blank=True)
ws.add_data_validation(dv_phase); dv_phase.add(f"F2:F{n+400}")
ws.add_data_validation(dv_status); dv_status.add(f"G2:G{n+400}")

# ------------------------------------------------------------------ Milestones
sheet("Milestones",
      ["MilestoneID", "ProjectID", "MilestoneName", "BaselineDate", "ForecastDate", "ActualDate", "Status"],
      D["milestones"], widths={"MilestoneName": 32},
      formats={"BaselineDate": DATE, "ForecastDate": DATE, "ActualDate": DATE})

# ------------------------------------------------------------------ Risks
RK_IN = ["RiskID", "ProjectID", "RiskTitle", "Category", "Probability", "Impact",
         "Response", "Owner", "Status", "DueDate"]
wsr = sheet("Risks", RK_IN + ["Exposure"], D["risks"], calc_from=len(RK_IN) + 1,
            widths={"RiskTitle": 52, "Category": 16, "Response": 12},
            formats={"DueDate": DATE})
for i in range(2, len(D["risks"]) + 2):
    wsr[f"K{i}"] = f"=E{i}*F{i}"
    wsr[f"K{i}"].font = CALC
    wsr[f"K{i}"].border = BORDER
dv_p = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
wsr.add_data_validation(dv_p); dv_p.add(f"E2:F{len(D['risks'])+400}")

# ------------------------------------------------------------------ Issues
sheet("Issues",
      ["IssueID", "ProjectID", "IssueTitle", "Severity", "Owner", "RaisedDate", "TargetDate", "Status"],
      D["issues"], widths={"IssueTitle": 52},
      formats={"RaisedDate": DATE, "TargetDate": DATE})

# ------------------------------------------------------------------ Lessons
sheet("LessonsLearned",
      ["LessonID", "ProjectID", "Category", "Lesson", "Recommendation", "DateCaptured"],
      D["lessons"], widths={"Lesson": 56, "Recommendation": 56},
      formats={"DateCaptured": DATE})

# ------------------------------------------------------------------ Config
# Values come from the dataset itself so the workbook and the dashboard can never drift apart.
NOTES = {
    "CPI_Green": "Cost Performance Index at or above this = Green",
    "CPI_Amber": "CPI at or above this (below Green) = Amber; below = Red",
    "SPI_Green": "Schedule Performance Index at or above this = Green",
    "SPI_Amber": "SPI at or above this (below Green) = Amber; below = Red",
    "Risk_Green": "Highest open risk exposure (Probability x Impact) at or below this = Green",
    "Risk_Amber": "Highest open risk exposure at or below this = Amber; above = Red "
                  "(so Red needs high probability AND high impact)",
    "Currency": "Reporting currency for all monetary values",
    "ReportingDate": "As-at date for elapsed and overdue calculations",
}
cfg_rows = [dict(Setting=k, Value=v, Notes=NOTES.get(k, "")) for k, v in D["config"].items()]
sheet("Config", ["Setting", "Value", "Notes"], cfg_rows, widths={"Setting": 18, "Notes": 62})

# ------------------------------------------------------------------ README
rd = wb.create_sheet("README", 0)
rd.column_dimensions["A"].width = 3
rd.column_dimensions["B"].width = 30
rd.column_dimensions["C"].width = 96
rows = [
    ("H1", "Portfolio Reporting - Data Template", ""),
    ("P", "", "Fill this workbook in and drop it onto the dashboard page. Every number on the dashboard is derived from these sheets."),
    ("H2", "Legend", ""),
    ("KV", "Blue header", "An INPUT column. Type your data here."),
    ("KV", "Grey header", "A CALCULATED column. Contains formulas - do not overwrite."),
    ("KV", "Sample rows", "Every sheet ships with realistic sample rows. Delete them and enter your own, or edit in place."),
    ("H2", "Hierarchy - how the drill-down is built", ""),
    ("KV", "Priorities", "Top level. Organisational priorities for the financial year."),
    ("KV", "Portfolios", "Each links to one Priority via PriorityID."),
    ("KV", "Programs", "Each links to one Portfolio via PortfolioID."),
    ("KV", "Projects", "Each links to one Program via ProgramID. This is the main sheet."),
    ("KV", "Milestones / Risks / Issues / LessonsLearned", "Each links to one Project via ProjectID."),
    ("P", "", "The only rule that matters: every ID referenced by a child row must exist in its parent sheet. Orphan rows are reported on the dashboard's Data Health panel and excluded from rollups."),
    ("H2", "The four numbers that drive everything", ""),
    ("KV", "BudgetAtCompletion (BAC)", "Total approved budget for the project."),
    ("KV", "PlannedValue (PV)", "Budgeted cost of work SCHEDULED to date."),
    ("KV", "EarnedValue (EV)", "Budgeted cost of work ACTUALLY COMPLETED to date. = BAC x % complete."),
    ("KV", "ActualCost (AC)", "What has actually been spent to date."),
    ("P", "", "From these four, the dashboard computes SPI, CPI, variances, forecast at completion, and every RAG status. If you only ever maintain four columns, maintain these."),
    ("H2", "PMI / PMBOK formulas used", ""),
    ("KV", "SV  = EV - PV", "Schedule Variance. Negative = behind schedule."),
    ("KV", "CV  = EV - AC", "Cost Variance. Negative = over budget."),
    ("KV", "SPI = EV / PV", "Schedule Performance Index. Below 1.0 = behind schedule."),
    ("KV", "CPI = EV / AC", "Cost Performance Index. Below 1.0 = over budget."),
    ("KV", "EAC = BAC / CPI", "Estimate At Completion - forecast total cost at current efficiency."),
    ("KV", "VAC = BAC - EAC", "Variance At Completion. Negative = forecast overrun."),
    ("KV", "Risk exposure = P x I", "Probability (1-5) x Impact (1-5), giving 1-25."),
    ("H2", "RAG rules", ""),
    ("P", "", "Thresholds live on the Config sheet - change them there, never in the dashboard code. Cost RAG uses CPI, Schedule RAG uses SPI, Risk RAG uses the highest open risk exposure on the project. Overall RAG is the worst of the three."),
    ("H2", "Assumptions in the sample data", ""),
    ("P", "", "All sample rows are synthetic and generated for demonstration only. Reporting date is set on the Config sheet. Currency is USD. No figure in this workbook is drawn from a real organisation."),
]
r = 1
for kind, k, v in rows:
    if kind == "H1":
        c = rd.cell(row=r, column=2, value=k); c.font = Font(name=ARIAL, size=16, bold=True); r += 2
    elif kind == "H2":
        r += 1
        c = rd.cell(row=r, column=2, value=k)
        c.font = Font(name=ARIAL, size=11, bold=True, color="1C5CAB"); r += 1
    elif kind == "KV":
        a = rd.cell(row=r, column=2, value=k); a.font = Font(name=ARIAL, size=10, bold=True)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = rd.cell(row=r, column=3, value=v); b.font = BODY
        b.alignment = Alignment(vertical="top", wrap_text=True); r += 1
    else:
        c = rd.cell(row=r, column=3, value=v); c.font = BODY
        c.alignment = Alignment(vertical="top", wrap_text=True); r += 1
rd.sheet_view.showGridLines = False

del wb["Sheet"]
wb.save("data/Portfolio_Reporting_Template.xlsx")
print("saved")
